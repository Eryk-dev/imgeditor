import os
import asyncio
import aiohttp
import openpyxl
from math import ceil
from datetime import timedelta
import logging
from aiohttp import ClientSession
from aiohttp import FormData
from asyncio import Semaphore
from typing import List, Optional
import ssl
import certifi  # Adicione esta importação

# Configurações iniciais
API_KEYS = [
    '2f5f1fd1956d9864edd9631da98fb335',
    '090cdf0ff97e053221c582049f1f413e',
    '51a49c9106fe87d90c569ade9918cee4',
    '8db752b75eb799c98d214c3390ee8bd7',
    '9539ab95c9fb7ddba6a7b34b6e890ef7'
]
EXPIRATION_TIME = 86400  # Tempo de expiração em segundos (1 semana)
SAVE_INTERVAL = 50  # Intervalo de salvamento em número de itens
CONCURRENT_UPLOADS = 10  # Número de uploads simultâneos
RETRY_LIMIT = 3  # Limite de tentativas de reintento
IMAGES_PER_KEY = 500  # Número de imagens por chave de API
PAUSE_DURATION = timedelta(minutes=0)  # Duração da pausa entre conjuntos de imagens

# Configuração de logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    handlers=[
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# Função para listar arquivos de imagem em um diretório local, incluindo subpastas
def list_local_files(directory: str) -> List[str]:
    image_extensions = ('.jpg', '.jpeg', '.png', '.bmp', '.gif', '.tiff', '.webp', '.heic', '.avif')
    image_files = []
    for root, _, files in os.walk(directory):
        for file in files:
            if file.lower().endswith(image_extensions):
                image_files.append(os.path.join(root, file))
    logger.info(f"Encontrados {len(image_files)} arquivos de imagem em '{directory}'.")
    return image_files

# Função auxiliar para ler o conteúdo do arquivo
def read_file(path: str) -> bytes:
    with open(path, 'rb') as f:
        return f.read()

# Função para fazer upload de imagens para ImgBB
async def upload_to_imgbb(session: ClientSession, image_path: str, api_key: str, retries: int = 0) -> Optional[dict]:
    try:
        # Ler o conteúdo do arquivo em um thread separado para não bloquear o loop de eventos
        file_content = await asyncio.to_thread(read_file, image_path)
        
        data = FormData()
        data.add_field('image', file_content, filename=os.path.basename(image_path))
        params = {'key': api_key, 'expiration': EXPIRATION_TIME}

        # Criar um contexto SSL usando certifi
        ssl_context = ssl.create_default_context(cafile=certifi.where())
        
        async with session.post("https://api.imgbb.com/1/upload", data=data, params=params, ssl=ssl_context) as response:
            if response.status != 200:
                logger.error(f"Erro no upload de '{image_path}': {response.status}")
                try:
                    error_text = await response.text()
                    logger.debug(f"Detalhes do erro: {error_text}")
                except Exception as e:
                    logger.debug(f"Erro ao obter detalhes do erro: {e}")
                return None
            try:
                json_response = await response.json()
                if json_response.get('success'):
                    return json_response
                else:
                    logger.error(f"Falha no upload de '{image_path}': {json_response}")
                    return None
            except aiohttp.ContentTypeError:
                logger.error("Erro: Resposta não está em formato JSON.")
                try:
                    error_text = await response.text()
                    logger.debug(f"Detalhes do erro: {error_text}")
                except Exception as e:
                    logger.debug(f"Erro ao obter detalhes do erro: {e}")
                return None
    except Exception as e:
        logger.error(f"Erro ao fazer upload de '{image_path}': {e}")
        if retries < RETRY_LIMIT:
            logger.info(f"Tentando novamente '{image_path}' (tentativa {retries + 1})")
            return await upload_to_imgbb(session, image_path, api_key, retries + 1)
        else:
            logger.error(f"Falha ao fazer upload de '{image_path}' após {RETRY_LIMIT} tentativas")
            return None

# Função para salvar resultados em um arquivo Excel
def save_to_excel(data: List[List], file_path: str):
    if not os.path.exists(file_path):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["Filename", "URL", "Display URL", "Delete URL", "Status"])
        logger.info(f"Arquivo Excel '{file_path}' criado com cabeçalho.")
    else:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

    for item in data:
        sheet.append(item)
    workbook.save(file_path)
    logger.info(f"Salvo {len(data)} registros no Excel '{file_path}'.")

# Função para ler o índice da última imagem processada da planilha Excel
def get_last_processed_index(file_path: str) -> int:
    if not os.path.exists(file_path):
        return 0
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    if sheet.max_row < 2:
        return 0  # Apenas cabeçalho presente
    return sheet.max_row - 1  # Subtrai 1 para ignorar o cabeçalho

# Função para solicitar novas chaves de API após o limite de 500 imagens
def request_new_api_keys():
    global API_KEYS  # Define que a variável API_KEYS é global para permitir modificá-la
    logger.warning("Limite de 500 imagens atingido para as chaves atuais.")
    new_api_keys = []
    for i in range(5):
        while True:
            new_key = input(f"Insira a nova chave API {i+1}: ").strip()
            if new_key:
                new_api_keys.append(new_key)
                break
            else:
                logger.warning("Chave API não pode estar vazia. Tente novamente.")
    # Confirmação das chaves inseridas
    logger.info("Você inseriu as seguintes chaves API:")
    for idx, key in enumerate(new_api_keys, 1):
        logger.info(f"API Key {idx}: {key}")
    confirm = input("As chaves acima estão corretas? (s/n): ").strip().lower()
    if confirm == 's':
        API_KEYS.clear()
        API_KEYS.extend(new_api_keys)
        logger.info("Novas chaves API atualizadas com sucesso.")
    else:
        logger.error("Chaves API não confirmadas. O processo será interrompido.")
        exit(1)

# Função para salvar resultados restantes ao finalizar o processamento
def finalize_excel_save(data: List[List], file_path: str):
    if data:
        save_to_excel(data, file_path)
        data.clear()

# Função principal para processar imagens em lote
async def process_batch(sem: Semaphore, session: ClientSession, files: List[str], start_index: int, output_excel: str, api_key: str, results: List[List]):
    async with sem:
        for i, file in enumerate(files):
            current_index = start_index + i + 1  # Atualiza o índice de progresso para refletir a posição real
            logger.info(f"Uploading {current_index}/{start_index + len(files)}: {file}")
            response = await upload_to_imgbb(session, file, api_key)
            if response:
                results.append([
                    os.path.basename(file),
                    response['data']['url'],
                    response['data']['display_url'],
                    response['data']['delete_url'],
                    response['status']
                ])
            # Salvar a cada SAVE_INTERVAL itens
            if len(results) >= SAVE_INTERVAL:
                save_to_excel(results, output_excel)
                results.clear()  # Limpa a lista de resultados após salvar

# Função para dividir a lista de arquivos em sublistas menores
def chunk_list(lst: List[str], n: int):
    if n <= 0:
        raise ValueError("O tamanho do chunk deve ser maior que zero.")
    for i in range(0, len(lst), n):
        yield lst[i:i + n]

# Função para gerenciar o processamento paralelo
async def process_images_parallel(directory: str, output_excel: str, concurrent_uploads: int):
    files = list_local_files(directory)
    total_files = len(files)
    last_processed_index = get_last_processed_index(file_path=output_excel)
    files_to_process = files[last_processed_index:]

    if not files_to_process:
        logger.info("Não há novos arquivos para processar.")
        return

    sem = Semaphore(concurrent_uploads)  # Limita o número de uploads simultâneos

    # Índice de chave API
    api_key_index = 0
    total_processed = 0
    results = []

    while total_processed < len(files_to_process):
        # Se chegarmos ao final da lista de chaves de API, solicitar novas
        if api_key_index >= len(API_KEYS):
            request_new_api_keys()
            api_key_index = 0  # Reinicia o índice para as novas chaves

        # Usar a chave de API atual
        api_key = API_KEYS[api_key_index]
        start_index = total_processed
        end_index = min(start_index + IMAGES_PER_KEY, len(files_to_process))
        files_chunk = files_to_process[start_index:end_index]

        if not files_chunk:
            break  # Se não houver mais arquivos para processar, encerra o loop

        # Garante que o tamanho do bloco não seja zero
        chunk_size = max(1, ceil(len(files_chunk) / concurrent_uploads))
        chunks = list(chunk_list(files_chunk, chunk_size))

        async with aiohttp.ClientSession() as session:
            tasks = [
                process_batch(sem, session, chunk, start_index + i * chunk_size, output_excel, api_key, results)
                for i, chunk in enumerate(chunks)
            ]
            await asyncio.gather(*tasks)

        # Salva os resultados restantes após cada chave de API
        finalize_excel_save(results, output_excel)

        total_processed += len(files_chunk)
        api_key_index += 1  # Incrementa o índice da chave de API

        # Pausa entre conjuntos de imagens, se necessário
        if PAUSE_DURATION:
            logger.info(f"Pausando por {PAUSE_DURATION} antes de continuar...")
            await asyncio.sleep(PAUSE_DURATION.total_seconds())

    # Salva qualquer resultado restante
    finalize_excel_save(results, output_excel)

    logger.info(f"Processamento concluído. Total de {total_processed} imagens enviadas.")

# Caminho do diretório local contendo as imagens
directory_path = '/Users/eryk/Library/CloudStorage/GoogleDrive-VENDASB2B@netair.com.br/O meu disco/NOVO LDRU FORMATADO'

# Caminho do diretório de documentos do usuário
documents_directory = '/Users/eryk/Downloads'
output_excel = os.path.join(documents_directory, 'novas fotos LDRU.xlsx')

# Chamada para processar imagens em paralelo
if __name__ == "__main__":
    try:
        asyncio.run(process_images_parallel(directory_path, output_excel, CONCURRENT_UPLOADS))
    except KeyboardInterrupt:
        logger.warning("Processo interrompido pelo usuário.")
    except Exception as e:
        logger.error(f"Ocorreu um erro inesperado: {e}")